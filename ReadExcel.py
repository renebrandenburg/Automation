import openpyxl
import os 


workbook = openpyxl.load_workbook("example.xlsx")
type(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')

cell = sheet['A1']
print(cell.value)
print(str(cell.value))

print(sheet['B2'].value)

print(sheet.cell(row=1,column=2))

for i in range(1,8):
    print(i,sheet.cell(row=i, column=2).value)